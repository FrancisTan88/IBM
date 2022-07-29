from selenium import webdriver  
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
from time import sleep
import numpy as np
import pandas as pd
import pyodbc
import requests
import sys



file_path = '/Users/kian199887/Downloads/github_francistan88/DSA/automatic_testing/submission_import.xlsx'
export_file = '/Users/kian199887/Downloads/github_francistan88/DSA/automatic_testing/submission_information_new拷貝4.xlsx'
df = pd.read_excel(file_path)
# book = load_workbook(export_file)
writer = pd.ExcelWriter(export_file, engine='openpyxl', date_format='yyyy-mm-dd')
# writer.book = book
# writer.sheets = {ws.title: ws for ws in book.worksheets}
# start_row = writer.sheets['工作表1'].max_row
ID_isnull = df['IdNo'].isnull()
k = founc()