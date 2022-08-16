from selenium import webdriver  
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
from time import sleep
import numpy as np
import pandas as pd
import pyodbc
import requests
import sys


"""
SC_Case 11, 13~15
"""

attribute_id = 'id'
attribute_class = 'class'
attribute_xpath = 'xpath'
attribute_css = 'css'


def LocateByAttribute(attribute, locate_name):
    if attribute == 'id':
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, locate_name)))

    elif attribute == 'class':
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, locate_name)))

    elif attribute == 'xpath':
        element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, locate_name)))
    
    elif attribute == 'css':
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, locate_name)))

    return element


# def LocateByText(locate_name, text_name):
#     text_name = '"' + text_name + '"'
#     element = driver.find_element(By.XPATH, locate_name + f'//*[contains(text(), {text_name})]')   
#     return element


def GetElementText(locate_name):
    # text = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, locate_name))).text
    text = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, locate_name))).get_attribute("value")
    return text


def Press(locate_name, attribute):
    element = LocateByAttribute(attribute, locate_name)
    element.click()
    return element


def Type(locate_name, type_value, attribute):
    element = LocateByAttribute(attribute, locate_name)
    element.send_keys(type_value)
    return element


def LogIn(login_email):
    locate_email = 'userEmail'
    try:
        element = Type(locate_email, login_email, attribute_id)
        time.sleep(1)
        element.send_keys(Keys.ENTER)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to log in")


def CreateCase(CaseType):
    try:
        ################################################################  Address  ################################################################################
        text_name = 'CBMY'

        locate_submission = '/html/body/app-root/div[1]/app-layout/div/app-side-menu/p-sidebar[2]/div/div/div/ul/li[1]/a/div'    
        # locate_company = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[1]/div/p-dropdown/div/span'
        locate_company = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[1]/div/p-dropdown/div/span'
        locate_CBMY = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[1]/div/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]'
        locate_NewApplication = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[2]/p-radiobutton/div'
        locate_ProductName = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[3]/div/div/p-dropdown/div/span'
        locate_space = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[1]/div[3]/div/div/p-dropdown/div/div[3]/div[1]/div/input'
        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-leading-page/div[2]/a'

        ################################################################  Execution: Create New Case  ################################################################################
        element = Press(locate_submission, attribute_xpath)
        time.sleep(2)
        # element = LocateByText(locate_company, text_name)
        element = Press(locate_company, attribute_xpath)
        # print(element.text)
        time.sleep(1)
        # element.click()
        element = Press(locate_CBMY, attribute_xpath)
        time.sleep(3)
        element = Press(locate_NewApplication, attribute_xpath)   
        time.sleep(1)
        element = Press(locate_ProductName, attribute_xpath)
        time.sleep(1) 
        element = Type(locate_space, CaseType, attribute_xpath)  
        time.sleep(1)
        element.send_keys(Keys.ARROW_DOWN)
        time.sleep(0.5)
        element.send_keys(Keys.ENTER)
        time.sleep(1)
        element = Press(locate_next, attribute_xpath)
        time.sleep(5)

    except:
        # driver.quit()
        sys.exit("fail to create case")


def FillCustomerInformation(ID_No):
    try:
        ################################################################  Address  ################################################################################
        locate_IDNO = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[1]/div[2]/input'
        locate_ResidentialStatus = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[3]/div/p-dropdown/div/span'
        locate_withoutlown = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[3]/div/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'
        locate_random = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[3]'
        locate_nationality = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[2]/div[4]/p-dropdown/div/div[2]/span'
        locate_citizen = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[3]/app-customer-information/div/form/div/div[2]/div[4]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        ################################################################  Execution: Fill it  ################################################################################
        element = Type(locate_IDNO, ID_No, attribute_xpath)
        time.sleep(1)
        element = Press(locate_random, attribute_xpath) 
        time.sleep(5)
        element = Press(locate_nationality, attribute_xpath)
        sleep(1)
        element = Press(locate_citizen, attribute_xpath)
        sleep(1)
        element = Press(locate_ResidentialStatus, attribute_xpath) 
        time.sleep(1)
        element = Press(locate_withoutlown, attribute_xpath)
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        element = Press(locate_next, attribute_xpath)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to fill in customer information")


def FillEmployment():
    try:
        ################################################################  Address  ################################################################################
        locate_occupation = '//*[@id="occupation"]/div/div[2]/span'
        locate_FactoryOperator = '//*[@id="occupation"]/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_MonthlyIncome = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[4]/app-employment/div/form/div[2]/div[2]/div[2]/p-inputnumber/span/input'
        locate_RegAdd = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[4]/app-employment/div/form/div[4]/div/div/app-address-input/form/div/div[3]/div[1]/div/button[1]/span'
        locate_WorkPhoneNo = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[4]/app-employment/div/form/div[5]/div[1]/input'
        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'

        MonthlyIncome = 88888
        WorkPhoneNo = 123

        ################################################################  Execution: Fill it  ################################################################################
        # element = Press(locate_occupation, attribute_xpath)
        # time.sleep(1)
        # element = Press(locate_FactoryOperator, attribute_xpath)
        # time.sleep(1)
        # element = Type(locate_MonthlyIncome, MonthlyIncome, attribute_xpath)
        # time.sleep(1)
        # element = Press(locate_RegAdd, attribute_xpath)
        # time.sleep(1)
        # element = Type(locate_WorkPhoneNo, WorkPhoneNo, attribute_xpath)
        # time.sleep(1)
        element = Press(locate_next, attribute_xpath)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to fill in employment")


def FillGuarantorPerson(PersonalID, CorporateID, CustomerName, MobilePhone):
    try:
        ################################################################  Address  ################################################################################
        locate_AddGuarantorPerson = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[2]/div/button/span[2]'
        locate_AddGuarantorPerson2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[2]/button/span[2]'

        locate_PersonalID = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[1]/div[2]/input'
        locate_CorporateID = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[1]/div[2]/input'
            
        locate_PersonalLegalRelationship = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[1]/div[3]/p-dropdown/div/div[2]'
        locate_CorporateLegalRelationship = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[1]/div[3]/p-dropdown/div/div[2]/span'
        locate_PersonalRelationship = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[2]/div[2]/p-dropdown/div/div[2]/span'
        locate_CorporateRelationship = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[2]/div[2]/p-dropdown/div/div[2]/span'
        locate_random = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[1]'

        locate_PersonalRace = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[3]/div[1]/p-dropdown/div/div[2]/span'
        locate_PersonalMalay = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[3]/div[1]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        locate_Brother = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[2]/div[2]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[4]/li'
        locate_Sister = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[2]/div[2]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[5]/li'

        locate_IdentityType = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[1]/div[1]/p-dropdown/div/div[2]/span'
        locate_IdentityCorporation = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[1]/div[1]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[2]/li'

        locate_PersonalGuarantor = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[1]/div[2]/form/div[1]/div[3]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_CorporateGuarantor = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[1]/div[3]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        locate_CustomerName = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[2]/div[1]/input'
        locate_MobilePhone = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[5]/app-guarantor-person/div[1]/div[2]/div[2]/form/div[2]/div[3]/input'

        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'

        ################################################################  Execution: Fill it  ################################################################################
        element = Press(locate_AddGuarantorPerson, attribute_xpath) 
        time.sleep(1)
        element = Press(locate_AddGuarantorPerson2, attribute_xpath) 
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        element = Type(locate_PersonalID, PersonalID, attribute_xpath)
        time.sleep(1)
        element = Press(locate_PersonalLegalRelationship, attribute_xpath)  #for the information loading
        time.sleep(3)
        element.click()
        time.sleep(1)
        element = Press(locate_PersonalGuarantor, attribute_xpath)
        # time.sleep(1)
        element = Press(locate_PersonalRelationship, attribute_xpath)
        time.sleep(1)
        element = Press(locate_Brother, attribute_xpath)
        # time.sleep(1)
        element = Press(locate_PersonalRace, attribute_xpath)
        time.sleep(1)
        element = Press(locate_PersonalMalay, attribute_xpath)
        

        element = Press(locate_IdentityType, attribute_xpath)
        time.sleep(1)
        element = Press(locate_IdentityCorporation, attribute_xpath)
        time.sleep(0.5)
        element = Type(locate_CorporateID, CorporateID, attribute_xpath)
        time.sleep(0.5)
        element = Press(locate_random, attribute_xpath)
        time.sleep(1)

        element = Press(locate_CorporateLegalRelationship, attribute_xpath)
        time.sleep(1)
        element = Press(locate_CorporateGuarantor, attribute_xpath)

        element = Press(locate_CorporateRelationship, attribute_xpath)
        time.sleep(1)
        element = Press(locate_Sister, attribute_xpath)

        text = GetElementText(locate_CustomerName)
        if not text: 
            element = Type(locate_CustomerName, CustomerName, attribute_xpath)
            time.sleep(1)
        
        element = Type(locate_MobilePhone, MobilePhone, attribute_xpath)
        time.sleep(1)

        element = Press(locate_next, attribute_xpath)
        time.sleep(3)
    
    except:
        # driver.quit()
        sys.exit("fail to fill in guarantor person")


def FillContactPerson():
    try:
        ################################################################  Address  ################################################################################
        locate_AddContactPerson = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div/div/button/span[2]'
        locate_AddContactPerson2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[2]/button/span[2]'

        locate_Name1 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[1]/form/div[1]/div[1]/input'
        locate_Name2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[2]/form/div[1]/div[1]/input'

        locate_ContactPersonRelationship1 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[1]/form/div[1]/div[2]/p-dropdown/div/div[2]/span'
        locate_ContactPersonRelationship2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[2]/form/div[1]/div[2]/p-dropdown/div/div[2]/span'

        locate_ContactPersonMobilePhone1 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[1]/form/div[2]/div[2]/input'
        locate_ContactPersonMobilePhone2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[2]/form/div[2]/div[2]/input'

        locate_parents = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[1]/form/div[1]/div[2]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_spouse = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[6]/app-contact-person/div[2]/form/div[1]/div[2]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[2]/li'

        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'

        Name1 = 'aa'
        Name2 = 'bb'
        Phone1 = '123'
        Phone2 = '321'

        ################################################################  Execution: Fill it  ################################################################################
        element = Press(locate_AddContactPerson, attribute_xpath)
        time.sleep(1)
        element = Press(locate_AddContactPerson2, attribute_xpath)
        time.sleep(1)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        element = Type(locate_Name1, Name1, attribute_xpath)
        
        element = Press(locate_ContactPersonRelationship1, attribute_xpath)
        time.sleep(1)
        element = Press(locate_parents, attribute_xpath)
        time.sleep(1)
        element = Type(locate_ContactPersonMobilePhone1, Phone1, attribute_xpath)
        time.sleep(1)

        element = Type(locate_Name2, Name2, attribute_xpath)
        time.sleep(1)
        element = Press(locate_ContactPersonRelationship2, attribute_xpath)
        time.sleep(1)
        element = Press(locate_spouse, attribute_xpath)
        time.sleep(1)
        element = Type(locate_ContactPersonMobilePhone2, Phone2, attribute_xpath)
        time.sleep(1)

        element = Press(locate_next, attribute_xpath)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to fill in contact person")
       


def FillCollateral():
    try:
        ################################################################  Address  ################################################################################
        locate_AddCollateral = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion'

        locate_property = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[1]/p-dropdown/div/div[2]'
        locate_FinanceAsset = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[1]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]'

        locate_category = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[2]/p-dropdown/div/div[2]/span'
        locate_CommercialVehicle = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[2]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        locate_HasValue = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[3]/p-dropdown/div/div[2]/span'
        locate_Y = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[3]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        locate_brand = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[4]/p-dropdown/div/div[2]/span'
        locate_Adiva = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[4]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[5]/li'

        locate_model = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[5]/p-dropdown/div/div[2]/span'
        locate_AD3 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[5]/p-dropdown/div/div[3]/div/ul/p-dropdownitem/li'

        locate_transaction = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[6]/p-dropdown/div/div[2]/span'
        locate_new = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[6]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'

        locate_date = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[7]/div[2]/p-calendar/span/button/span[1]'
        locate_left = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[7]/div[2]/p-calendar/span/div/div[1]/div/div/button[1]'
        locate_August = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[7]/div[2]/p-calendar/span/div/div[2]/span[8]'

        locate_manu = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[10]/p-dropdown/div/div[2]/span'
        locate_BMW = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[10]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li/span[1]'

        locate_transmission = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[11]/p-dropdown/div/div[2]/span'
        locate_auto = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[11]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[2]/li'

        locate_PurchasePrice = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[15]/div[2]/sigv-currency/div/p-inputnumber/span/input'
        locate_SalesApprisalPrice = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[1]/div[16]/div[2]/sigv-currency/div/p-inputnumber/span/input'

        locate_DownPayment = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[2]/div[1]/div[2]/sigv-currency/div/p-inputnumber/span/input'

        locate_GPSinstallation = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[2]/div[3]/p-dropdown/div/div[2]/span'
        locate_None = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[7]/app-collateral/div[1]/div/p-accordion/div/p-accordiontab/div/div[2]/div/div[1]/div[2]/app-collateral-vehicle-motor/div/form/div/div[2]/div[3]/p-dropdown/div/div[3]/div/ul/p-dropdownitem/li'

        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'

        PurchasePrice = '99999'
        SalesApprisalPrice = PurchasePrice
        DownPayment = '33333'

        ################################################################  Execution: Fill it  ################################################################################
        element = Press(locate_AddCollateral, attribute_xpath)
        time.sleep(10)

        element = Press(locate_property, attribute_xpath)
        time.sleep(1)
        element = Press(locate_FinanceAsset, attribute_xpath)

        element = Press(locate_category, attribute_xpath)
        time.sleep(1)
        element = Press(locate_CommercialVehicle, attribute_xpath)

        element = Press(locate_HasValue, attribute_xpath)
        time.sleep(1)
        element = Press(locate_Y, attribute_xpath)

        element = Press(locate_brand, attribute_xpath)
        time.sleep(1.5)
        element = Press(locate_Adiva, attribute_xpath)
        time.sleep(3)   

        element = Press(locate_model, attribute_xpath)
        time.sleep(1)
        element = Press(locate_AD3, attribute_xpath)

        element = Press(locate_transaction, attribute_xpath)
        time.sleep(1)
        element = Press(locate_new, attribute_xpath)

        element = Press(locate_date, attribute_xpath)
        time.sleep(1)
        element = Press(locate_left, attribute_xpath)
        time.sleep(1)
        element = Press(locate_August, attribute_xpath)

        element = Press(locate_manu, attribute_xpath)
        time.sleep(1)
        element = Press(locate_BMW, attribute_xpath)

        element = Press(locate_transmission, attribute_xpath)
        time.sleep(1)
        element = Press(locate_auto, attribute_xpath)

        element = Type(locate_PurchasePrice, PurchasePrice, attribute_xpath)
        time.sleep(1)
        element = Type(locate_SalesApprisalPrice, SalesApprisalPrice, attribute_xpath)
        time.sleep(1)
        element = Type(locate_DownPayment, DownPayment, attribute_xpath)
        time.sleep(1)

        element = Press(locate_GPSinstallation, attribute_xpath)
        time.sleep(1)
        element = Press(locate_None, attribute_xpath)
        time.sleep(1)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        element = Press(locate_next, attribute_xpath)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to fill in collateral")
        

def FillTermsConditions():
    try:
        ################################################################  Address  ################################################################################
        locate_DealSource = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[1]/p-dropdown/div/div[2]/span'
        locate_CarDealer = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[1]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_DealerName = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[2]/p-dropdown/div/div[2]/span'
        locate_InputName = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[2]/p-dropdown/div/div[3]/div[1]/div/input'
        locate_SalesName = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[3]/p-dropdown/div/div[2]/span'
        locate_PAIDAIAH = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[2]/div[3]/p-dropdown/div/div[3]/div/ul/p-dropdownitem[1]/li'
        locate_QuotesType = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[3]/div[1]/p-dropdown/div/div[2]/span'
        locate_ETP = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[3]/div[1]/p-dropdown/div/div[3]/div/ul/p-dropdownitem/li'
        locate_InterestRate = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[1]/div[3]/div[2]/div[2]/p-inputnumber/span/input'
        locate_CommisionDeduction = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[2]/div/div/p-table/div/div/table/tbody/tr[1]/td[5]/p-checkbox/div/div[2]'
        locate_ApplyTerms = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[4]/div/div[1]/p-inputnumber/span/input'
        locate_random2 = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[8]/app-terms-conditions/div/div[4]'
        locate_next = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/div[2]/div/div[10]/div[2]/a'

        DealerName = 'JHR0001 Twit'
        InterestRate = 8
        Terms = 50

        ################################################################  Execution: Fill it  ################################################################################
        element = Press(locate_DealSource, attribute_xpath)
        time.sleep(1)
        element = Press(locate_CarDealer, attribute_xpath)

        element = Press(locate_DealerName, attribute_xpath)
        time.sleep(1)
        element = Type(locate_InputName, DealerName, attribute_xpath)
        time.sleep(1)
        element.send_keys(Keys.ARROW_DOWN)
        time.sleep(0.5)
        element.send_keys(Keys.ENTER)
        time.sleep(1)

        element = Press(locate_SalesName, attribute_xpath)
        time.sleep(1)
        element = Press(locate_PAIDAIAH, attribute_xpath)
        time.sleep(0.5)

        element = Press(locate_QuotesType, attribute_xpath)
        time.sleep(1)
        element = Press(locate_ETP, attribute_xpath)
        time.sleep(0.5)

        element = Type(locate_InterestRate, InterestRate, attribute_xpath)
        time.sleep(1)

        element = Press(locate_CommisionDeduction, attribute_xpath)
        time.sleep(1)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        element = Type(locate_ApplyTerms, Terms, attribute_xpath)
        time.sleep(1)

        element = Press(locate_random2, attribute_xpath)
        time.sleep(3)

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)

        element = Press(locate_next, attribute_xpath)
        time.sleep(3)

    except:
        # driver.quit()
        sys.exit("fail to fill in terms and conditions")
        

def FillAttachment():
    try:
        locate_submit = '/html/body/app-root/div[1]/app-layout/div/div/div/app-process/sigv-fixed-bottom-panel/div/div/div/div/div/div[1]/a'

        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        element = Press(locate_submit, attribute_xpath)
        # time.sleep(15)

    except:
        # driver.quit()
        sys.exit("fail to press the submit button")


def CheckAlert():
    try:
        locate_popup = '/html/body/app-root/div[1]/app-layout/p-toast/div/p-toastitem/div/div/div//*[contains(text(), "Submit Success")]'
        text_name = 'Submit Success'
        element = LocateByAttribute(attribute_xpath, locate_popup)
        sleep(1)
        if text_name in element.text:
            print(element.text)
            print("The case is submitted successfully", '\n')
        else:
            print("Not  success", '\n')
    
    except:
        # driver.quit()
        sys.exit("The case is failed")


def GetEmailFromJson(obj):
    return obj['email']


def GetSqlData(cursor, index, by):
    if by == 'IdNo':
        SQL_data = cursor.execute(f"""select CaseNo, CreateTime, CurrentApplicantId, StatusID, StageId, ApplyDate, IdNo, ProductCode from CreditRatingScales.Submission
                        where IdNo = '{index}' 
                        ORDER BY CreateTime DESC
                        """).fetchone()
    elif by == 'CaseNo':
        SQL_data = cursor.execute(f"""select CaseNo, CreateTime, CurrentApplicantId, StatusID, StageId, ApplyDate, IdNo, ProductCode from CreditRatingScales.Submission
                        where CaseNo = '{index}' 
                        ORDER BY CreateTime DESC
                        """).fetchone()

    sleep(2)
    CaseNo = SQL_data[0]
    CurrentApplicantId = SQL_data[2]
    StatusID = SQL_data[3]
    return CaseNo, CurrentApplicantId, StatusID


def GetApplicantEmail(url):
    response = requests.get(url)
    json_data = response.json()
    email = GetEmailFromJson(json_data)
    return email


def DataPreprocessing(df, id_isnull):
    for i in range(len(df['IdNo'])):
        if id_isnull[i] == False:
            df['IdNo'].iloc[i] = str(df['IdNo'].iloc[i]).replace('.0', '')
            df['Guarantor Person(Indi)'].iloc[i] = str(df['Guarantor Person(Indi)'].iloc[i]).replace('.0', '')
            df['Guarantor Person(Corpo)'].iloc[i] = str(df['Guarantor Person(Corpo)'].iloc[i]).replace(',', '')
            df['Mobile Phone'].iloc[i] = str(df['Mobile Phone'].iloc[i]).replace('.0', '')
            df['Mobile Phone'].iloc[i] = '0' + str(df['Mobile Phone'].iloc[i])

    return df


# def AddCaseToDF(df, row_data, dic_column, CaseNo, status):
#     dic = {}
#     date = datetime.today().strftime('%Y-%m-%d')

#     for i in range(len(dic_column)):
#         dic[dic_column[str(i)]] = row_data[i]

#     df = df.append(dic, ignore_index=True)
#     df['CaseNo'].iloc[len(df)-1] = CaseNo
#     df['Date'].iloc[len(df)-1] = date
#     df['Status'].iloc[len(df)-1] = status
#     return df


def RowDataToDF(row_no, dataframe, dic_column):
    date = datetime.today().strftime('%Y-%m-%d')
    row_data = dataframe.iloc[row_no]
    num_column = dataframe.shape[1]
    dictionary = {}

    for i in range(num_column):
        dictionary[dic_column[str(i)]] = row_data[i]

    df_row = pd.DataFrame(dictionary, index=[0])
    return df_row




if __name__ == "__main__":

    ############################################################ Submission ####################################################################################
    # read excel and convert to dataframe
    file_path = '/Users/kian199887/Downloads/github_francistan88/DSA/automatic_testing/submission_import.xlsx'
    export_file = '/Users/kian199887/Downloads/github_francistan88/DSA/automatic_testing/submission_export.xlsx'
    df = pd.read_excel(file_path)
    book = load_workbook(export_file)
    writer = pd.ExcelWriter(export_file, engine='openpyxl', date_format='yyyy-mm-dd')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    start_row = writer.sheets['工作表1'].max_row
    ID_isnull = df['IdNo'].isnull()
    column_name = {
        '0': 'Date',
        '1': 'CaseNo',
        '2': 'IdNo',
        '3': 'Guarantor Person(Indi)',
        '4': 'Guarantor Person(Corpo)',
        '5': 'Customer Name',
        '6': 'Mobile Phone',
        '7': 'CIF No(Corp)',
        '8': 'Status',
        '9': 'Product Name'
    } 
    login_email = 'nabiladibidris@chailease.com.my'
    CaseType = 'SC_Case'

    # remove'.0' for all and add '0' in front of the Mobile Phone
    df = DataPreprocessing(df, ID_isnull)

    # Get row data:
    print("\n")
    RowNo = int(input("Please input the row number from the excel file submission_import.xlsx: "))
    while (RowNo < 5 and RowNo != 3 ) or (RowNo > 7 ):
        print("The case data can not be applied to SC0011 or SC0013 ~ SC0015")
        RowNo = int(input("Please input another row number from the excel file submission_import.xlsx: "))
    df_row = RowDataToDF(RowNo, df, column_name)
    ID_No = int(df['IdNo'].iloc[RowNo])
    PersonalID = int(df['Guarantor Person(Indi)'].iloc[RowNo])
    CorporateID = df['Guarantor Person(Corpo)'].iloc[RowNo]
    CustomerName = df['Customer Name'].iloc[RowNo] 
    MobilePhone = df['Mobile Phone'].iloc[RowNo]
    
    # open the submission web 
    s = Service('./chromedriver')
    driver = webdriver.Chrome(service=s)
    driver.maximize_window()
    url = 'https://sit01-websubmission.chailease.com.my/websubmission-ui/'
    driver.get(url)
    time.sleep(3)

    # Submit
    LogIn(login_email)    

    CreateCase(CaseType)

    FillCustomerInformation(ID_No)

    FillEmployment()

    FillGuarantorPerson(PersonalID, CorporateID, CustomerName, MobilePhone)

    FillContactPerson()

    FillCollateral()

    FillTermsConditions()

    FillAttachment()

    CheckAlert()


    ############################################################# Preliminary Credit Review ####################################################################################
    # connect to SQL server
    server = 'Please input your server name'
    database = 'Your DataBase'
    username = 'Please input your username'
    password = 'Please input your password'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 18 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)
    cursor = cnxn.cursor()
    
    # get CaseNo, ApplicantId
    by = 'IdNo'
    CaseNo, CurrentApplicantId, StatusId = GetSqlData(cursor, ID_No, by)
    sleep(1)

    date = datetime.today().strftime('%Y-%m-%d')
    df_row['Date'].loc[0] = date
    df_row['CaseNo'].loc[0] = CaseNo
    df_row['Status'].loc[0] = StatusId
    
    # add the case data to the excel file
    print('\n')
    print(df_row)
    print('\n')
    df_row.to_excel(writer, sheet_name='工作表1', startrow=start_row, index=False, header=False)
    writer.save()

    # get credit officer's Email through API
    api_url = f"http://10.164.55.100:8000/dev-backdoor/system-management/Rbac/UserProfile/{CurrentApplicantId}"
    OfficerEmail = GetApplicantEmail(api_url)
    print(OfficerEmail)
    print('\n')



